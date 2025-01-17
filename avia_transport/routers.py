from models import *
from flask import render_template, request

from functools import wraps
from flask import session, redirect, url_for
from openpyxl import Workbook
from flask import Response
from models import Flight, Aviacompany, Ticket, Airplane, Client
from config import app

#сохранение в pdf
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from flask import Response
from io import BytesIO

@app.route('/')
def index():
    return render_template("index.html")
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        role = request.form.get('role', 2)

        if Users.select().where(Users.username == username).exists():
            return "Username already exists", 400

        user = Users(username=username, role=role)
        user.set_password(password)
        user.save()

        return redirect('/login')

    return render_template('register.html')
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        user = Users.get_or_none(Users.username == username)

        if user and user.check_password(password):
            session['user_id'] = user.id
            session['role'] = user.role

            if session['role'] == 1:
                return redirect('/admin_dashboard')
            else:
                return redirect('/dashboard')

        return "Invalid username or password", 400

    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('user_id', None)
    session.pop('role', None)
    return redirect('/login')

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'role' not in session or session['role'] != 1:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/flights')
def get_flights():
    flights = Flight.select()
    return render_template("flights.html", flights=flights)


@app.route('/aviacompanies')
def get_aviacompanies():
    aviacompanies = Aviacompany.select()
    return render_template("aviacompanies.html", aviacompanies=aviacompanies)


@app.route('/tickets')
def get_tickets():
    tickets = Ticket.select()
    return render_template("tickets.html", tickets=tickets)


@app.route('/airplanes')
def get_airplanes():
    airplanes = Airplane.select()
    return render_template("airplanes.html", airplanes=airplanes)


@app.route('/clients', methods=['GET'])
def get_clients():
    aviacompanies = Aviacompany.select()
    clients = Client.select()
    return render_template('client.html', client=clients, aviacompanies=aviacompanies)



@app.route('/add_client', methods=['POST'])
@admin_required
def add_client():
    id = request.form['id']
    name = request.form['name']
    phone_number = request.form['phone_number']
    flight_hours = request.form.get('flight_hours') or None
    luggage = request.form.get('luggage') or None
    aviacompany_id = request.form.get('aviacompany_id')

    if not aviacompany_id:
        return "Aviacompany ID is required!", 400

    try:
        aviacompany = Aviacompany.get(Aviacompany.id == aviacompany_id)
    except Aviacompany.DoesNotExist:
        return f"Aviacompany with id {aviacompany_id} does not exist!", 400

    Client.create(
        id = id,
        name=name,
        phone_number=phone_number,
        flight_hours=flight_hours,
        luggage=luggage,
        aviacom_id=aviacompany_id
    )

    return redirect('/clients')



@app.route('/delete_client/<int:client_id>', methods=['POST'])
@admin_required
def delete_client(client_id):
    client = Client.get(Client.id == client_id)
    client.delete_instance()
    return redirect('/clients')

@app.route('/update_client/<int:client_id>', methods=['GET', 'POST'])
@admin_required
def update_client(client_id):
    client = Client.get(Client.id == client_id)

    if request.method == 'GET':
        return render_template('update_client.html', client=client)

    else:
        client.name = request.form['name']
        client.phone_number = request.form['phone_number']
        client.flight_hours = request.form.get('flight_hours') or None
        client.luggage = request.form.get('luggage') or None
        client.aviacompany_id = request.form['aviacompany_id']
        client.save()
        return redirect('/clients')

@app.route('/add_flight', methods=['POST'])
@admin_required
def add_flight():
    id = request.form['id']
    depature_point = request.form['depature_point']
    arrival_point = request.form['arrival_point']
    depature_time = request.form['depature_time']
    arrival_time = request.form['arrival_time']

    Flight.create(
        id=id,
        depature_point=depature_point,
        arrival_point=arrival_point,
        depature_time=depature_time,
        arrival_time=arrival_time
    )

    return redirect('/flights')

@app.route('/delete_flight/<int:flight_id>', methods=['POST'])
@admin_required
def delete_flight(flight_id):
    flight = Flight.get(Flight.id == flight_id)
    flight.delete_instance()
    return redirect('/flights')

@app.route('/update_flight/<int:flight_id>', methods=['GET', 'POST'])
@admin_required
def update_flight(flight_id):
    flight = Flight.get(Flight.id == flight_id)

    if request.method == 'GET':
        return render_template('update_flight.html', flight=flight)

    else:
        flight.depature_point = request.form['depature_point']
        flight.arrival_point = request.form['arrival_point']
        flight.depature_time = request.form['depature_time']
        flight.arrival_time = request.form['arrival_time']
        flight.save()
        return redirect('/flights')


@app.route('/delete_aviacompany/<int:aviacompany_id>', methods=['POST'])
@admin_required
def delete_aviacompany(aviacompany_id):
    company = Aviacompany.get(Aviacompany.id == aviacompany_id)
    company.delete_instance()
    return redirect('/aviacompanies')

@app.route('/update_aviacompany/<int:aviacompany_id>', methods=['GET', 'POST'])
@admin_required
def update_aviacompany(aviacompany_id):
    company = Aviacompany.get(Aviacompany.id == aviacompany_id)

    if request.method == 'GET':
        return render_template('update_aviacompany.html', company=company)

    else:
        company.name = request.form['name']
        company.planes_amount = request.form['planes_amount']
        company.save()
        return redirect('/aviacompanies')

@app.route('/add_aviacompany', methods=['POST'])
@admin_required
def add_aviacompany():
    id = request.form['id']
    name = request.form['name']
    planes_amount = request.form['planes_amount']

    Aviacompany.create(
        id=id,
        name=name,
        planes_amount=planes_amount
    )

    return redirect('/aviacompanies')

@app.route('/add_ticket', methods=['POST'])
@admin_required
def add_ticket():
    id = request.form['id']
    cost = request.form['cost']
    landing_class = request.form['landing_class']
    flight_id = request.form['flight_id']

    try:
        flight = Flight.get(Flight.id == flight_id)
    except Flight.DoesNotExist:
        return f"Flight with id {flight_id} does not exist!", 400

    Ticket.create(
        id=id,
        cost=cost,
        landing_class=landing_class,
        flight_id=flight_id
    )

    return redirect('/tickets')

@app.route('/update_ticket/<int:ticket_id>', methods=['GET', 'POST'])
@admin_required
def update_ticket(ticket_id):
    ticket = Ticket.get(Ticket.id == ticket_id)

    if request.method == 'GET':
        flights = Flight.select()
        return render_template('update_ticket.html', ticket=ticket, flights=flights)

    else:
        ticket.cost = request.form['cost']
        ticket.landing_class = request.form['landing_class']
        ticket.flight_id = request.form['flight_id']

        ticket.save()

        return redirect('/tickets')

@app.route('/delete_ticket/<int:ticket_id>', methods=['POST'])
@admin_required
def delete_ticket(ticket_id):
    ticket = Ticket.get(Ticket.id == ticket_id)
    ticket.delete_instance()
    return redirect('/tickets')

@app.route('/add_airplane', methods=['POST'])
@admin_required
def add_airplane():
    id = request.form['id']
    business_seats = request.form['business_seats']
    econom_seats = request.form['econom_seats']
    luggage_capacity = request.form['luggage_capacity']
    aviacompany_id = request.form['aviacompany_id']
    flight_id = request.form['flight_id']

    Airplane.create(
        id=id,
        business_seats=business_seats,
        econom_seats=econom_seats,
        luggage_capacity=luggage_capacity,
        aviacompany_id=aviacompany_id,
        flight_id=flight_id
    )

    return redirect('/airplanes')

@app.route('/delete_airplane/<int:airplane_id>', methods=['POST'])
@admin_required
def delete_airplane(airplane_id):
    airplane = Airplane.get(Airplane.id == airplane_id)

    airplane.delete_instance()

    return redirect('/airplanes')

@app.route('/update_airplane/<int:airplane_id>', methods=['GET', 'POST'])
@admin_required
def update_airplane(airplane_id):
    airplane = Airplane.get(Airplane.id == airplane_id)

    if request.method == 'GET':
        return render_template('update_airplane.html', airplane=airplane)

    else:
        airplane.business_seats = request.form['business_seats']
        airplane.econom_seats = request.form['econom_seats']
        airplane.luggage_capacity = request.form['luggage_capacity']
        airplane.aviacompany_id = request.form['aviacompany_id']
        airplane.flight_id = request.form['flight_id']
        airplane.save()

        return redirect('/airplanes')

@app.route('/dashboard')
def dashboard():
    if session['role'] == 2:
        return render_template('dashboard.html')
    return redirect('/logout')

@app.route('/admin_dashboard')
def admin_dashboard():
    if session['role'] == 1:
        return render_template('admin_dashboard.html')
    return redirect('/logout')

@app.route('/flights_view')
def get_flights_view():
    flights = Flight.select()
    return render_template("flights_view.html", flights=flights)


@app.route('/aviacompanies_view')
def get_aviacompanies_view():
    aviacompanies = Aviacompany.select()
    return render_template("aviacompanies_view.html", aviacompanies=aviacompanies)


@app.route('/tickets_view')
def get_tickets_view():
    tickets = Ticket.select()
    return render_template("tickets_view.html", tickets=tickets)


@app.route('/airplanes_view')
def get_airplanes_view():
    airplanes = Airplane.select()
    return render_template("airplanes_view.html", airplanes=airplanes)


@app.route('/clients_view', methods=['GET'])
def get_clients_view():
    aviacompanies = Aviacompany.select()
    clients = Client.select()
    return render_template('client_view.html', client=clients, aviacompanies=aviacompanies)

@app.route('/export/xlsx')
@admin_required
def export_xlsx():

    wb = Workbook()


    ws_flight = wb.active
    ws_flight.title = "Flights"
    ws_flight.append(['ID', 'Departure Point', 'Arrival Point', 'Departure Time', 'Arrival Time'])

    flights = Flight.select()
    for flight in flights:
        ws_flight.append(
            [flight.id, flight.depature_point, flight.arrival_point, flight.depature_time, flight.arrival_time])


    ws_aviacompany = wb.create_sheet(title="Aviacompanies")
    ws_aviacompany.append(['ID', 'Name', 'Planes Amount'])

    aviacompanies = Aviacompany.select()
    for aviacompany in aviacompanies:
        ws_aviacompany.append([aviacompany.id, aviacompany.name, aviacompany.planes_amount])


    ws_ticket = wb.create_sheet(title="Tickets")
    ws_ticket.append(['ID', 'Cost', 'Landing Class', 'Flight ID'])

    tickets = Ticket.select()
    for ticket in tickets:

        ws_ticket.append([ticket.id, ticket.cost, ticket.landing_class, ticket.flight_id.id])


    ws_airplane = wb.create_sheet(title="Airplanes")
    ws_airplane.append(['ID', 'Business Seats', 'Econom Seats', 'Luggage Capacity', 'Aviacompany ID', 'Flight ID'])

    airplanes = Airplane.select()
    for airplane in airplanes:

        ws_airplane.append([airplane.id, airplane.business_seats, airplane.econom_seats, airplane.luggage_capacity,
                            airplane.aviacompany_id.id, airplane.flight_id.id])  # Здесь исправление

    ws_client = wb.create_sheet(title="Clients")
    ws_client.append(['ID', 'Name', 'Phone Number', 'Flight Hours', 'Luggage', 'Aviacompany ID'])

    clients = Client.select()
    for client in clients:
        ws_client.append(
            [client.id, client.name, client.phone_number, client.flight_hours, client.luggage, client.aviacom_id.id])  # Здесь исправление

    ws_users = wb.create_sheet(title="Users")
    ws_users.append(['ID', 'Username', 'Password', 'Role'])

    users = Users.select()
    for user in users:
        ws_users.append([user.id, user.username, user.password, user.role])

    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={"Content-Disposition": "attachment;filename=database_export.xlsx"})


@app.route('/export/pdf')
@admin_required
def export_pdf():
    buffer = BytesIO()

    doc = SimpleDocTemplate(buffer, pagesize=letter)

    styles = getSampleStyleSheet()

    content = []

    def add_table_title(title):
        paragraph = Paragraph(title, styles['Heading2'])
        content.append(paragraph)

    flights = Flight.select()
    flight_data = [['ID', 'Departure Point', 'Arrival Point', 'Departure Time', 'Arrival Time']]
    for flight in flights:
        flight_data.append(
            [flight.id, flight.depature_point, flight.arrival_point, flight.depature_time, flight.arrival_time])

    add_table_title("Flight Data")
    flight_table = Table(flight_data)
    flight_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    content.append(flight_table)

    aviacompanies = Aviacompany.select()
    aviacompany_data = [['ID', 'Name', 'Planes Amount']]
    for aviacompany in aviacompanies:
        aviacompany_data.append([aviacompany.id, aviacompany.name, aviacompany.planes_amount])

    add_table_title("Aviacompany Data")
    aviacompany_table = Table(aviacompany_data)
    aviacompany_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    content.append(aviacompany_table)

    tickets = Ticket.select()
    ticket_data = [['ID', 'Cost', 'Landing Class', 'Flight ID']]
    for ticket in tickets:
        ticket_data.append([ticket.id, ticket.cost, ticket.landing_class, ticket.flight_id.id])

    add_table_title("Ticket Data")
    ticket_table = Table(ticket_data)
    ticket_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    content.append(ticket_table)

    airplanes = Airplane.select()
    airplane_data = [['ID', 'Business Seats', 'Econom Seats', 'Luggage Capacity', 'Aviacompany ID', 'Flight ID']]
    for airplane in airplanes:
        airplane_data.append([airplane.id, airplane.business_seats, airplane.econom_seats, airplane.luggage_capacity,
                              airplane.aviacompany_id.id, airplane.flight_id.id])

    add_table_title("Airplane Data")
    airplane_table = Table(airplane_data)
    airplane_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    content.append(airplane_table)

    clients = Client.select()
    client_data = [['ID', 'Name', 'Phone Number', 'Flight Hours', 'Luggage', 'Aviacompany ID']]
    for client in clients:
        client_data.append(
            [client.id, client.name, client.phone_number, client.flight_hours, client.luggage, client.aviacom_id.id])

    add_table_title("Client Data")
    client_table = Table(client_data)
    client_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    content.append(client_table)

    users = Users.select()
    user_data = [['ID', 'Username', 'Password (Shortened)', 'Role']]
    for user in users:
        shortened_password = user.password[:20] + "..." if len(user.password) > 20 else user.password
        user_data.append([user.id, user.username, shortened_password, user.role])

    add_table_title("Users Data")
    user_table = Table(user_data)
    user_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    content.append(user_table)

    doc.build(content)

    buffer.seek(0)
    return Response(buffer, mimetype='application/pdf',
                    headers={"Content-Disposition": "attachment;filename=database_export.pdf"})